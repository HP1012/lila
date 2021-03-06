# -*- coding: utf-8 -*-

import io
import logging
import sys
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_testlog(file):
    '''Parse testlog file'''
    try:
        logger.debug("Parse testlog %s", Path(file).name)
        lst = ['func_full', 'src_full', 'c0', 'c1', 'mcdc', 'test_time']
        with open(file, encoding='shift-jis', errors='ignore') as fp:
            lines = [l.strip() for l in fp.readlines()[:len(lst)]]

        data = {lst[i]: lines[i][lines[i].index(':')+1:].strip()
                for i in range(len(lst))}

        parts = Path(data['src_full']).parts

        data.update({
            'src_rel': '/'.join(parts),
            'src_name': parts[-1],
            'func': Path(data.get('func_full')).name,
            'src_short': '/'.join(parts[-3:])
        })
    except Exception as e:
        logger.exception(e)
        data = {}
    finally:
        return data


def get_xlsx_raw(xlsx, sheet, begin=1, end=sys.maxsize, headers={}):
    '''Get raw data of table from excel.'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get raw data from %s %s", Path(xlsx).name, sheet)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]

        data = [[val(cell) for cell in row] for row in wb[sheet].rows]
        data = data[begin-1:min(end, sys.maxsize)]

        first_row = data[0][:]
        data[0] = [headers.get(col, col) for col in data[0]]

        if headers != {}:
            data.append(first_row)

    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data


def get_xlsx_cells(xlsx, sheet, list_cell):
    '''Get cell value from excel file'''
    def val(cell):
        return str(cell.value) if cell.is_date else cell.value

    logger.debug("Get value of cell %s", list_cell)

    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        sheet = sheet if isinstance(sheet, str) else wb.sheetnames[sheet-1]
        ws = wb[sheet]

        data = {key: val(ws[key]) for key in list_cell}
    except Exception as e:
        logger.exception(e)
        data = {}
    finally:
        wb.close()
        return data


def get_xlsx_sheets(xlsx):
    '''Get sheets of xlsx'''
    logger.debug("Get sheets from file %s", xlsx)
    try:
        with open(xlsx, 'rb') as fp:
            xlsx = io.BytesIO(fp.read())

        wb = load_workbook(xlsx, read_only=True)
        data = wb.sheetnames
    except Exception as e:
        logger.exception(e)
        data = []
    finally:
        wb.close()
        return data
